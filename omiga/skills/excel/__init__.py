"""Excel processing skill for Omiga."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from omiga.skills.base import Skill, SkillContext, SkillMetadata, SkillError

logger = logging.getLogger("omiga.skills.excel")


class ExcelSkill(Skill):
    """Skill for processing Excel files."""

    metadata = SkillMetadata(
        name="excel",
        description="Excel 文件处理 - 读取、写入、修改 Excel 文件",
        emoji="📊",
        tags=["excel", "xlsx", "spreadsheet", "data"],
    )

    def __init__(self, context: SkillContext):
        super().__init__(context)

    async def execute(
        self,
        action: str,
        file_path: Optional[str] = None,
        output_path: Optional[str] = None,
        sheet_name: Optional[str] = None,
        data: Optional[List[List[Any]]] = None,
        **kwargs: Any,
    ) -> Any:  # type: ignore[override]
        """Execute the Excel skill.

        Args:
            action: Action to perform (read, write, append, info, list_sheets)
            file_path: Input file path
            output_path: Output file path
            sheet_name: Sheet name to operate on
            data: Data to write (list of rows)
            **kwargs: Action-specific arguments

        Returns:
            Result of the Excel operation
        """
        actions = {
            "read": self._read_excel,
            "write": self._write_excel,
            "append": self._append_excel,
            "info": self._get_excel_info,
            "list_sheets": self._list_sheets,
            "get_sheet_data": self._get_sheet_data,
        }

        if action not in actions:
            raise SkillError(f"Unknown action: {action}", self.name)

        return await actions[action](
            file_path=file_path,
            output_path=output_path,
            sheet_name=sheet_name,
            data=data,
            **kwargs,
        )

    async def _read_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        header_row: int = 0,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Read data from an Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise SkillError(
                "openpyxl not installed. Run: pip install openpyxl", self.name
            )

        path = Path(file_path)
        if not path.exists():
            raise SkillError(f"File not found: {file_path}", self.name)

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise SkillError(f"Sheet '{sheet_name}' not found", self.name)
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            rows = list(ws.iter_rows(values_only=True))
            headers = rows[header_row] if rows else []
            data_rows = rows[header_row + 1:] if len(rows) > header_row else []

            wb.close()

            return {
                "file": str(path),
                "sheet": sheet_name,
                "headers": list(headers) if headers else [],
                "row_count": len(data_rows),
                "data": data_rows[:100],  # Limit to first 100 rows
            }
        except Exception as e:
            raise SkillError(f"Failed to read Excel: {e}", self.name)

    async def _write_excel(
        self,
        output_path: str,
        data: List[List[Any]],
        sheet_name: str = "Sheet1",
        headers: Optional[List[str]] = None,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Write data to a new Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise SkillError(
                "openpyxl not installed. Run: pip install openpyxl", self.name
            )

        if not data:
            raise SkillError("No data to write", self.name)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            row_num = 1
            if headers:
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num, value=header)
                row_num += 1

            for row in data:
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                row_num += 1

            output = Path(output_path)
            output.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output))
            wb.close()

            return {
                "status": "success",
                "output": str(output),
                "sheet": sheet_name,
                "rows_written": len(data),
            }
        except Exception as e:
            raise SkillError(f"Failed to write Excel: {e}", self.name)

    async def _append_excel(
        self,
        file_path: str,
        data: List[List[Any]],
        sheet_name: Optional[str] = None,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Append data to an existing Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise SkillError(
                "openpyxl not installed. Run: pip install openpyxl", self.name
            )

        path = Path(file_path)
        if not path.exists():
            raise SkillError(f"File not found: {file_path}", self.name)

        try:
            wb = openpyxl.load_workbook(str(path))

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise SkillError(f"Sheet '{sheet_name}' not found", self.name)
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            start_row = ws.max_row + 1
            for row in data:
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=start_row, column=col_num, value=value)
                start_row += 1

            wb.save(str(path))
            wb.close()

            return {
                "status": "success",
                "file": str(path),
                "sheet": sheet_name,
                "rows_appended": len(data),
            }
        except Exception as e:
            raise SkillError(f"Failed to append to Excel: {e}", self.name)

    async def _get_excel_info(
        self,
        file_path: str,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Get Excel file metadata."""
        try:
            import openpyxl
        except ImportError:
            raise SkillError(
                "openpyxl not installed. Run: pip install openpyxl", self.name
            )

        path = Path(file_path)
        if not path.exists():
            raise SkillError(f"File not found: {file_path}", self.name)

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

            sheets_info = []
            for name in wb.sheetnames:
                ws = wb[name]
                sheets_info.append({
                    "name": name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                })

            file_size = path.stat().st_size
            wb.close()

            return {
                "file": str(path),
                "sheet_count": len(wb.sheetnames),
                "sheets": sheets_info,
                "file_size_bytes": file_size,
            }
        except Exception as e:
            raise SkillError(f"Failed to get Excel info: {e}", self.name)

    async def _list_sheets(
        self,
        file_path: str,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """List all sheets in an Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise SkillError(
                "openpyxl not installed. Run: pip install openpyxl", self.name
            )

        path = Path(file_path)
        if not path.exists():
            raise SkillError(f"File not found: {file_path}", self.name)

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True)
            sheets = list(wb.sheetnames)
            wb.close()

            return {"file": str(path), "sheets": sheets}
        except Exception as e:
            raise SkillError(f"Failed to list sheets: {e}", self.name)

    async def _get_sheet_data(
        self,
        file_path: str,
        sheet_name: str,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Get all data from a specific sheet."""
        result = await self._read_excel(
            file_path=file_path,
            sheet_name=sheet_name,
        )
        return result

    def get_usage(self) -> str:
        """Return usage instructions."""
        return """
Excel Skill - Excel 文件处理

可用操作:
- read <file_path> [sheet_name]: 读取 Excel 数据
- write <output_path> <data> [headers]: 写入新 Excel
- append <file_path> <data> [sheet_name]: 追加数据
- info <file_path>: 获取文件信息
- list_sheets <file_path>: 列出所有工作表
- get_sheet_data <file_path> <sheet_name>: 获取工作表数据

示例:
- 读取 Excel: execute(action="read", file_path="data.xlsx")
- 写入 Excel: execute(action="write", output_path="out.xlsx", data=[["a", "b"], [1, 2]])
- 列出工作表：execute(action="list_sheets", file_path="data.xlsx")

需要安装：pip install openpyxl
"""
